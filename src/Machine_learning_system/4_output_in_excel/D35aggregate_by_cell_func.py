# %%
from D31output_in_excel import store_title, filename, output_df
import D32tidy_and_output
import D33other_data_output
import D34output_in_red

import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Font


wb = openpyxl.load_workbook(filename)
ws = wb[store_title]

font_color = '008080'

cell = ws.cell(7, 7)
cell.value = '集計'
cell.font = Font(bold=True, color=font_color)

cell_value_dict = {
    'データ総額': f'=SUM(D4:D{ws.max_row})',
    ' 内 決済完了額': f'=SUMIF(F4:F{ws.max_row},"<>キャンセル",D4:D{ws.max_row})',
    ' 内 キャンセル額': f'=SUMIF(F4:F{ws.max_row},"=キャンセル",D4:D{ws.max_row})'}

row = 8
col = 8

for title, func in cell_value_dict.items():
    cell = ws.cell(row, 8)
    cell.value = title
    if row == 8:
        cell.font = Font(bold=True, color=font_color)
    else:
        cell.font = Font(bold=True)

    cell = ws.cell(row, 10)
    cell.value = func

    row += 1

filename = f'./output/{store_title}.xlsx'
wb.save(filename)
wb.close()

# %%
