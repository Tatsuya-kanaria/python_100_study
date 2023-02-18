# %%
from D31output_in_excel import store_df, store_title, filename
import D32tidy_and_output

import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Font


def calc_delta(t):
    t1, t2 = t
    delta = t2 - t1
    return delta.total_seconds() / 60


store_df['order_accept_datetime'] = pd.to_datetime(
    store_df['order_accept_date'])

store_df['delivered_datetime'] = pd.to_datetime(store_df['delivered_date'])

store_df['delta'] = store_df[
    ['order_accept_datetime',
     'delivered_datetime'
     ]].apply(calc_delta, axis=1)

delivery_time = store_df.groupby(['store_id'])['delta'].describe()
delivery_time

wb = openpyxl.load_workbook(filename)
ws = wb[store_title]

cell = ws.cell(1, 7)
cell.value = f'配達完了までの時間'
cell.font = Font(bold=True, color='008080')

rows = dataframe_to_rows(delivery_time, index=False, header=True)

# 表の貼り付け位置
row_start = 3
col_start = 8

side = Side(style='thin', color='008080')
border = Border(top=side, bottom=side, left=side, right=side)

for row_no, row in enumerate(rows, row_start):
    for col_no, value in enumerate(row, col_start):
        cell = ws.cell(row_no, col_no)
        cell.value = value
        cell.border = border
        if row_no == row_start:
            cell.fill = PatternFill(patternType='solid', fgColor='008080')
            cell.font = Font(bold=True, color='FFFFFF')

filename = f'./output/{store_title}.xlsx'
wb.save(filename)
wb.close()

# %%
