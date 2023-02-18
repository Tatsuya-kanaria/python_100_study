# %%
from D31output_in_excel import filename, store_title

import openpyxl
# スタイル関係のインポート
from openpyxl.styles import PatternFill, Border, Side, Font


wb = openpyxl.load_workbook(filename)
ws = wb[store_title]

side = Side(style='thin', color='008080')
border = Border(top=side, bottom=side, left=side, right=side)

# データの表に部分に罫線を設定
for row in ws:
    for cell in row:
        if ws[cell.coordinate].value:
            ws[cell.coordinate].border = border

ws.cell(1, 1).font = Font(bold=True, color='008080')
# タイトル行の指定
row = 3
# 指定
fill_patternType = 'solid'
fill_fgColor = '008080'
# フォント色
font_color = 'FFFFFF'

cell_value_dict = {2: '注文日時', 3: '顧客ID', 4: '購入総額', 5: '注文タイプ', 6: '注文状態'}

for col, value in cell_value_dict.items():
    cell = ws.cell(row, col)
    cell.fill = PatternFill(patternType=fill_patternType, fgColor=fill_fgColor)
    cell.value = value
    cell.font = Font(bold=True, color=font_color)

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 12

# ファイルに保存
wb.save(filename)
wb.close()

# %%
