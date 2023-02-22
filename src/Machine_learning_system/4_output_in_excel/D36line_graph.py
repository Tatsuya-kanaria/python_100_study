# %%
from D31output_in_excel import store_title, filename, output_df
import D32tidy_and_output
import D33other_data_output
import D34output_in_red
import D35aggregate_by_cell_func

import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Font
from openpyxl.chart import Reference, BarChart, PieChart, LineChart, ScatterChart, Series


wb = openpyxl.load_workbook(filename)
ws = wb[store_title]

cell = ws.cell(7, 7)
cell.value = f'売上グラフ'
cell.font = Font(bold=True, color='008080')

# グラフ用の参照データを指定、D列（購入総額）の4行目から20件を指定
refy = Reference(ws, min_col=4, min_row=4, max_col=4, max_row=23)

# グラフシリーズを生成
series = Series(refy, title='売上額')

# Chart
chart = LineChart()
chart.title = '折れ線グラフ'
chart.x_axis.title = '件数'
chart.y_axis.title = '売上額'
chart.height = 10
chart.width = 20
chart.series.append(series)

# 生成したChartオブジェクトをシートの指定位置に追加
ws.add_chart(chart, 'H12')

filename = f'./output/{store_title}.xlsx'
wb.save(filename)
wb.close()

# %%
