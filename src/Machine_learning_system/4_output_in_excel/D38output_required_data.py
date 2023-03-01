# %%
from D31output_in_excel import filename, store_id, store_df
from D37report_preparation import cancel_rank, get_area_rank_df, make_store_daily, get_area_delivery_rank_df

from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl
import os


# 最初にテスト用のファイルを削除
if os.path.exists('./output/test.xlsx'):
    os.remove('./output/test.xlsx')
if os.path.exists(filename):
    os.remove(filename)


def data_sheet_output(trg_wb, sheetname, target_df, indexFlg):
    ws = trg_wb.create_sheet(title=sheetname)
    rows = dataframe_to_rows(target_df, index=indexFlg, header=True)

    # 表の貼り付け位置
    row_start = 1
    col_start = 1

    for row_no, row in enumerate(rows, row_start):
        for col_no, value in enumerate(row, col_start):
            ws.cell(row_no, col_no).value = value

    # データシートは非表示にしておく
    ws.sheet_state = 'hidden'


def make_data_sheet(trg_id, trg_st_df, targetfolder):
    target_daily = make_store_daily(trg_id)
    store_name = trg_st_df['store_name'].unique()[0]

    # 新たにファイルを作成する
    store_title = f'{trg_id}_{store_name}'

    wb = openpyxl.Workbook()

    # キャンセルランキング
    data_sheet_output(
        wb,
        'Data_CancelRank',
        cancel_rank,
        False
    )

    # エリア売上ランキング
    data_sheet_output(
        wb,
        'Data_AreaRank', get_area_rank_df(trg_id),
        False
    )

    # エリア配達時間ランキング
    data_sheet_output(
        wb,
        'Data_DeliveryRank',
        get_area_delivery_rank_df(trg_id),
        False
    )

    # 該当店舗の日単位売上データ
    data_sheet_output(
        wb,
        'Data_Target_Daily',
        target_daily,
        False
    )

    filename = os.path.join(targetfolder, f'./output/{store_title}.xlsx')

    wb.save(filename)
    wb.close()

    return filename


filename_store = make_data_sheet(store_id, store_df, '')

# %%
