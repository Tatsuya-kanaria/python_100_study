# %%
from D31output_in_excel import m_store, store_id, store_name
from D37report_preparation import cancel_rank, check_store_cancel_rank, check_store_sales_rank, check_store_delivery_rank
from D38output_required_data import filename_store

from openpyxl.styles import Font
from openpyxl.chart import Reference, LineChart, Series
import openpyxl


def make_summary_sheet(trg_id, storename, trgfile):
    target_cancel_rank = check_store_cancel_rank(trg_id)
    target_sales_rank = check_store_sales_rank(trg_id)
    target_delivery_rank = check_store_delivery_rank(trg_id)

    wb = openpyxl.load_workbook(trgfile)
    ws = wb.active
    ws.title = 'サマリーレポート'

    cell = ws.cell(1, 1)
    cell.value = f'{storename} サマリーレポート（4月〜6月）'
    cell.font = Font(bold=True, color='008080', size=20)

    # 売上ランキングの表示
    tmpWs = wb['Data_Target_Daily']
    cell = ws.cell(3, 2)
    cell.value = '店舗売上額'
    cell.font = Font(bold=True, color='008080', size=16)

    # セルの結合
    ws.merge_cells('E3:F3')

    cell = ws.cell(3, 5)
    cell.value = f'=SUM({tmpWs.title}!B2:B{tmpWs.max_row})'
    cell.font = Font(bold=True, color='0080FF', size=16)
    cell.number_format = '#,##0'

    cell = ws.cell(4, 2)
    cell.value = '店舗売上ランク'
    cell.font = Font(bold=True, color='008080', size=16)

    cell = ws.cell(4, 5)
    cell.value = f'{len(m_store)}店舗中 {target_sales_rank} 位'
    cell.font = Font(bold=True, color='0080FF', size=16)

    # グラフ用の参照データを指定
    refy = Reference(tmpWs, min_col=2, min_row=2,
                     max_col=2, max_row=tmpWs.max_row)

    # グラフシリーズを生成
    series = Series(refy, title='売上額')

    # Chart
    chart = LineChart()
    chart.title = '期間売上額（日毎）'
    chart.x_axis.title = '件数'
    chart.y_axis.title = '売上額'
    chart.height = 10
    chart.width = 15
    chart.series.append(series)

    # 生成したChartオフジェクトをシートの指定位置に追加
    ws.add_chart(chart, 'B6')

    # 地域情報
    tmpWs = wb['Data_AreaRank']

    cell = ws.cell(4, 10)
    cell.value = '地域店舗売上情報'
    cell.font = Font(bold=True, color='008080', size=16)

    cell = ws.cell(5, 11)
    cell.value = '最高額'

    cell = ws.cell(5, 12)
    cell.value = f'=MAX({tmpWs.title}!B2:B{tmpWs.max_row})'
    cell.number_format = '#,##0'

    cell = ws.cell(6, 11)
    cell.value = '最低額'

    cell = ws.cell(6, 12)
    cell.value = f'=MIN({tmpWs.title}!B2:B{tmpWs.max_row})'
    cell.number_format = '#,##0'

    cell = ws.cell(7, 11)
    cell.value = '地域平均'

    cell = ws.cell(7, 12)
    cell.value = f'=AVERAGE({tmpWs.title}!B2:B{tmpWs.max_row})'
    cell.number_format = '#,##0'

    # キャンセル率の表示
    cell = ws.cell(11, 10)
    cell.value = 'キャンセルランク'
    cell.font = Font(bold=True, color='008080', size=16)

    cell = ws.cell(12, 11)
    cell.value = f'{len(m_store)}店舗中 {target_cancel_rank} 位'
    cell.font = Font(bold=True, color='008080', size=16)

    tmpWs = wb['Data_CancelRank']

    cell = ws.cell(13, 11)
    cell.value = '地域平均'

    cell = ws.cell(13, 12)
    cell.value = f'=AVERAGE({tmpWs.title}!C2:C{tmpWs.max_row})'
    cell.number_format = '0.00'

    # 配達時間ランキングの表示
    cell = ws.cell(15, 10)
    cell.value = '配達時間ランク'
    cell.font = Font(bold=True, color='008080', size=16)

    cell = ws.cell(16, 11)
    cell.value = f'{len(m_store)}店舗中 {target_delivery_rank} 位'
    cell.font = Font(bold=True, color='008080', size=16)

    tmpWs = wb['Data_DeliveryRank']

    cell = ws.cell(17, 11)
    cell.value = '地域平均'

    cell = ws.cell(17, 12)
    cell.value = f'=AVERAGE({tmpWs.title}!B2:B{tmpWs.max_row})'
    cell.number_format = '0.00'

    wb.save(trgfile)
    wb.close


make_summary_sheet(store_id, store_name, filename_store)

# %%
