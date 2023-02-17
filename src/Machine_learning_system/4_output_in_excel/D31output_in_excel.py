# %%
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl
import pandas as pd
import glob
import os


m_store = pd.read_csv('./data/m_store.csv')
m_area = pd.read_csv('./data/m_area.csv')

current_dir = os.getcwd()
tbl_order_file = os.path.join(current_dir, './data/tbl_order_*.csv')
tbl_order_files = glob.glob(tbl_order_file)

order_all = pd.DataFrame()
for file in tbl_order_files:
    order_tmp = pd.read_csv(file)
    print(f'{file}: {len(order_tmp)}')
    order_all = pd.concat([order_all, order_tmp], ignore_index=True)

# 保守用店舗データの削除
order_all = order_all.loc[order_all['store_id'] != 999]

order_all = pd.merge(order_all, m_store, on='store_id', how='left')
order_all = pd.merge(order_all, m_area, on='area_cd', how='left')

# マスターにないコードに対応した文字列を設定
takeout_name_dict = {
    0: 'デリバリー',
    1: 'お持ち帰り',
}

status_name_dict = {
    0: '受付',
    1: 'お支払済み',
    2: 'お渡し済み',
    9: 'キャンセル'
}


def set_str(dict, flag_name):
    for flag, name in dict.items():
        order_all.loc[order_all[flag_name] == flag,
                      flag_name.replace('_flag', '') + '_name'] = name


set_str(takeout_name_dict, 'takeout_flag')
set_str(status_name_dict, 'status')

order_all['order_date'] = pd.to_datetime(
    order_all['order_accept_date']).dt.date

order_all.head()

if __name__ == '__main__':
    # pyxl test
    wb = openpyxl.Workbook()
    ws = wb['Sheet']
    ws.cell(1, 1).value = '書き込みテストです。'
    wb.save('./output/test.xlsx')
    wb.close()

    wb = openpyxl.load_workbook('./output/test.xlsx', read_only=True)
    ws = wb['Sheet']
    print(ws.cell(1, 1).value)
    wb.close()

# テストデータの準備
store_id = 1
store_df = order_all.loc[order_all['store_id'] == store_id].copy()
store_name = store_df['store_name'].unique()[0]

store_sales_total = store_df.loc[store_df['status'].isin(
    [1, 2])]['total_amount'].sum()
store_sales_takeout = store_df.loc[
    store_df['status'] == 1]['total_amount'].sum()
store_sales_delivery = store_df.loc[
    store_df['status']
    == 2]['total_amount'].sum()
print(f'売上額確認 {store_sales_total} = {store_sales_takeout + store_sales_delivery}')
output_df = store_df[['order_accept_date', 'customer_id',
                      'total_amount', 'takeout_name', 'status_name']]
output_df.head()

store_title = f'{store_id}_{store_name}'

wb = openpyxl.Workbook()
ws = wb.active
ws.title = store_title

ws.cell(1, 1).value = f'{store_title} 売上データ'

# OpenPyXLのユーティリティdataframe_to_rowsを利用
rows = dataframe_to_rows(output_df, index=False, header=True)

# 表の貼り付け位置
row_start = 3
col_start = 2

for row_no, row in enumerate(rows, row_start):
    for col_no, value in enumerate(row, col_start):
        ws.cell(row_no, col_no).value = value

filename = f'./output/{store_title}.xlsx'
wb.save(filename)
wb.close()

# %%
