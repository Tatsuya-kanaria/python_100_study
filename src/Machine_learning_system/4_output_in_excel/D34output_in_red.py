# %%
from D31output_in_excel import store_title, filename, output_df
import D32tidy_and_output
import D33other_data_output

import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Font


wb = openpyxl.load_workbook(filename)
ws = wb[store_title]

rows = dataframe_to_rows(output_df, index=False, header=True)

# 表の貼り付け位置
row_start = 3
col_start = 2

for row_no, row in enumerate(rows, row_start):
    if row_no == row_start:
        continue
    for col_no, value in enumerate(row, col_start):
        ws.cell(row_no, col_no).value = value
        if value == 'キャンセル':
            ws.cell(row_no, col_no).font = Font(bold=True, color='FF0000')

filename = f'./output/{store_title}.xlsx'
wb.save(filename)
wb.close()

# %%
