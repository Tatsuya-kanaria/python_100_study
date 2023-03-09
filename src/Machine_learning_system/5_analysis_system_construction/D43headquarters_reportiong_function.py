# %%
from D42check_mechanism import m_store, max_str_date, target_data
from D41basic_folder_generation import output_dir

import pandas as pd
import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Font


def get_rank_df(target_data):
    # 店舗のデータ作成、ランキングDFの返却
    tmp = target_data.loc[target_data['status'].isin([1, 2])]
    rank = tmp.groupby(['store_id'])[
        'total_amount'].sum().sort_values(ascending=False)
    rank = pd.merge(rank, m_store, on='store_id', how='left')

    return rank


def get_cancel_rank_df(target_data):
    # キャンセル率の計算、ランキングDFの返却
    cancel_df = pd.DataFrame()
    cancel_cnt = target_data.loc[target_data['status'] == 9].groupby(['store_id'])[
        'store_id'].count()
    order_cnt = target_data.loc[target_data['status'].isin(
        [1, 2, 9])].groupby(['store_id'])['store_id'].count()
    cancel_rate = (cancel_cnt / order_cnt) * 100
    cancel_df['cancel_rate'] = cancel_rate
    cancel_df = pd.merge(cancel_df, m_store, on='store_id', how='left')
    cancel_df = cancel_df.sort_values('cancel_rate', ascending=True)

    return cancel_df


def data_export(df, ws, row_start, col_start):
    # スタイル定義
    side = Side(style='thin', color='008080')
    border = Border(top=side, bottom=side, left=side, right=side)

    rows = dataframe_to_rows(df, index=False, header=True)

    for row_no, row in enumerate(rows, row_start):
        for col_no, value in enumerate(row, col_start):
            cell = ws.cell(row_no, col_no)
            cell.value = value
            cell.border = border
            if row_no == row_start:
                cell.fill = PatternFill(patternType='solid', fgColor='008080')
                cell.font = Font(bold=True, color='FFFFFF')

# 本部向けレポーティングデータ処理


def make_report_hq(target_data, output_folder):
    rank = get_rank_df(target_data)
    cancel_rank = get_cancel_rank_df(target_data)

    # Excel出力処理
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'サマリーレポート（本部向け）'

    cell = ws.cell(1, 1)
    cell.value = f'本部向け {max_str_date}月度 サマリーレポート'
    cell.font = Font(bold=True, color='008080', size=20)

    cell = ws.cell(3, 2)
    cell.value = f'{max_str_date}月度 サマリーレポート'
    cell.font = Font(bold=True, color='008080', size=20)

    cell = ws.cell(3, 6)
    cell.value = f"{'{:,}'.format(rank['total_amount'].sum())}"
    cell.font = Font(bold=True, color='008080', size=20)

    # 売上ランキングを直接出力
    cell = ws.cell(5, 2)
    cell.value = f'売上ランキング'
    cell.font = Font(bold=True, color='008080', size=16)

    # 表の貼り付け
    data_export(rank, ws, 6, 2)

    # キャンセル率ランキングを直接出力
    cell = ws.cell(5, 8)
    cell.value = f'キャンセル率ランキング'
    cell.font = Font(bold=True, color='008080', size=16)

    # 表の貼り付け位置
    data_export(cancel_rank, ws, 6, 8)

    wb.save(os.path.join(output_folder, f'report_hq_{max_str_date}.xlsx'))
    wb.close()


# make_report_hq(target_data, output_dir)

# %%
