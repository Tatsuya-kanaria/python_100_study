# %%
from D41basic_folder_generation import output_dir
from D42check_mechanism import m_store, max_str_date, target_data
from D43headquarters_reportiong_function import get_rank_df, get_cancel_rank_df, data_export

import os
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Font


def get_store_rank(target_id, target_df):
    rank = get_rank_df(target_df)
    store_rank = rank.loc[rank['store_id'] == target_id].index + 1

    return store_rank[0]


def get_store_sale(target_id, target_df):
    rank = get_rank_df(target_df)
    store_sale = rank.loc[rank['store_id'] == target_id]['total_amount']

    return store_sale


def get_store_cancel_rank(target_id, target_df):
    cancel_df = get_cancel_rank_df(target_df)
    cancel_df = cancel_df.reset_index()
    store_cancel_rank = cancel_df.loc[
        cancel_df['store_id'] == target_id].index + 1

    return store_cancel_rank[0]


def get_store_cancel_count(target_id, target_df):
    store_cancel_count = target_df.loc[(target_df['status'] == 9) & (
        target_df['store_id'] == target_id)].groupby(['store_id'])['store_id'].count()

    return store_cancel_count


def get_delivery_rank_df(target_id, target_df):
    delivery = target_df.loc[target_df['status'] == 2]
    delivery_rank = delivery.groupby(
        ['store_id'])['delta'].mean().sort_values()
    delivery_rank = pd.merge(delivery_rank, m_store, on='store_id', how='left')

    return delivery_rank


def get_delivery_rank_store(target_id, target_df):
    delivery_rank = get_delivery_rank_df(target_id, target_df)
    store_delivery_rank = delivery_rank.loc[delivery_rank['store_id']
                                            == target_id].index + 1

    return store_delivery_rank[0]

# 店舗向けレポーティングデータ処理


def make_report_store(target_data, target_id, output_folder):
    rank = get_store_rank(target_id, target_data)
    sale = get_store_sale(target_id, target_data)
    cancel_rank = get_store_cancel_rank(target_id, target_data)
    cancel_count = get_store_cancel_count(target_id, target_data)
    delivery_df = get_delivery_rank_df(target_id, target_data)
    delivery_rank = get_delivery_rank_store(target_id, target_data)

    store_name = m_store.loc[
        m_store['store_id'] == target_id]['store_name'].values[0]

    # Excel出力処理
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '店舗向けレポーティング'

    cell = ws.cell(1, 1)
    cell.value = f'{store_name} {max_str_date}月度 サマリーレポート'
    cell.font = Font(bold=True, color='008080', size=20)

    cell = ws.cell(3, 2)
    cell.value = f'{max_str_date}月度 売上総額'
    cell.font = Font(bold=True, color='008080', size=20)

    cell = ws.cell(3, 6)
    cell.value = f"{'{:,}'.format(sale.values[0])}"
    cell.font = Font(bold=True, color='008080', size=20)

    # 売上ランキングを直接出力
    cell = ws.cell(5, 2)
    cell.value = f'売上ランキング'
    cell.font = Font(bold=True, color='008080', size=16)

    cell = ws.cell(5, 5)
    cell.value = f'{rank}位'
    cell.font = Font(bold=True, color='008080', size=16)

    cell = ws.cell(6, 2)
    cell.value = f'売上データ'
    cell.font = Font(bold=True, color='008080', size=16)

    # 表の貼り付け
    tmp_df = target_data.loc[(target_data['store_id'] == target_id) & (
        target_data['status'].isin([1, 2]))]
    tmp_df = tmp_df[['order_accept_date', 'customer_id',
                     'total_amount', 'takeout_name', 'status_name']]
    data_export(tmp_df, ws, 7, 2)

    # 配達完了までの時間を直接出力
    ave_time = delivery_df.loc[delivery_df['store_id']
                               == target_id]['delta'].values[0]
    cell = ws.cell(5, 14)
    cell.value = f'配達完了までの時間ランキング'
    cell.font = Font(bold=True, color='008080', size=16)

    cell = ws.cell(5, 18)
    cell.value = f'{delivery_rank}位 平均{ave_time}分'
    cell.font = Font(bold=True, color='008080', size=16)

    cell = ws.cell(6, 14)
    cell.value = f'各店舗の配達時間ランク'
    cell.font = Font(bold=True, color='008080', size=16)

    # 表の貼り付け
    data_export(delivery_df, ws, 7, 14)

    wb.save(os.path.join(output_folder,
            f'{target_id}_{store_name}_report_{max_str_date}.xlsx'))
    wb.close()


# store_id = 20
# make_report_store(target_data, store_id, output_dir)


# %%
