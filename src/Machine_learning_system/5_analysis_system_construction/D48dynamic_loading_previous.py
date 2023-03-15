# %%
from D41basic_folder_generation import output_dir
from D42check_mechanism import m_store
from D43headquarters_reportiong_function import get_rank_df, get_cancel_rank_df, data_export
from D44stores_reporting_function import get_store_rank, get_store_sale, get_store_cancel_rank, get_store_cancel_count, get_delivery_rank_df, get_delivery_rank_store


from openpyxl.styles import PatternFill, Border, Side, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl
import pandas as pd
import os


# 本部向けレポーティングデータ処理（過去月データ対応Ver）


def make_report_hq_r2(target_data_list, output_folder):
    # Excel出力処理
    wb = openpyxl.Workbook()

    file_date = ''

    for tmp in target_data_list:
        df = pd.DataFrame(tmp)

        df_date = pd.to_datetime(df['order_accept_date']).max()
        trg_date = df_date.strftime('%Y%m')

        if file_date == '':
            # 初回のみファイル名用に年月を保持
            file_date = trg_date

        rank = get_rank_df(df)
        cancel_rank = get_cancel_rank_df(df)

        # ワークシートの作成
        ws = wb.create_sheet(title=f'{trg_date}月度')

        cell = ws.cell(1, 1)
        cell.value = f'本部向け {trg_date}月度 サマリーレポート'
        cell.font = Font(bold=True, color='008080', size=20)

        cell = ws.cell(3, 2)
        cell.value = f'{trg_date}月度 サマリーレポート'
        cell.font = Font(bold=True, color='008080', size=20)

        cell = ws.cell(3, 6)
        cell.value = f"{'{:,}'.format(rank['total_amount'].sum())}"
        cell.font = Font(bold=True, color='008080', size=20)

        # 売上ランキングを直接出力
        cell = ws.cell(5, 2)
        cell.value = f'売上ランキング'
        cell.font = Font(bold=True, color='008080', size=16)

        # 表の貼り付け
        data_export(rank, ws, 6, 2)

        # キャンセル率ランキングを直接出力
        cell = ws.cell(5, 8)
        cell.value = f'キャンセル率ランキング'
        cell.font = Font(bold=True, color='008080', size=16)

        # 表の貼り付け位置
        data_export(cancel_rank, ws, 6, 8)

    # デフォルトシートは削除
    wb.remove(wb.worksheets[0])

    # DFループが終わったらブックを保存
    wb.save(os.path.join(output_folder, f'report_hq_{trg_date}.xlsx'))
    wb.close()

# 店舗向けレポーティングデータ処理


def make_report_store_r2(target_data_list, target_id, output_folder):
    # Excel出力処理
    wb = openpyxl.Workbook()

    file_date = ''

    for tmp in target_data_list:
        df = pd.DataFrame(tmp)

        df_date = pd.to_datetime(df['order_accept_date']).max()
        trg_date = df_date.strftime('%Y%m')

        if file_date == '':
            # 初回のみファイル名用に年月を保持
            file_date = trg_date

        rank = get_store_rank(target_id, df)
        sale = get_store_sale(target_id, df)
        cancel_rank = get_store_cancel_rank(target_id, df)
        cancel_count = get_store_cancel_count(target_id, df)
        delivery_df = get_delivery_rank_df(target_id, df)
        delivery_rank = get_delivery_rank_store(target_id, df)

        store_name = m_store.loc[
            m_store['store_id'] == target_id]['store_name'].values[0]

        # ワークシート作成
        ws = wb.create_sheet(title=f'{trg_date}月度')

        # Excel出力処理
        cell = ws.cell(1, 1)
        cell.value = f'{store_name} {trg_date}月度 サマリーレポート'
        cell.font = Font(bold=True, color='008080', size=20)

        cell = ws.cell(3, 2)
        cell.value = f'{trg_date}月度 売上総額'
        cell.font = Font(bold=True, color='008080', size=20)

        cell = ws.cell(3, 6)
        cell.value = f"{'{:,}'.format(sale.values[0])}"
        cell.font = Font(bold=True, color='008080', size=20)

        # 売上ランキングを直接出力
        cell = ws.cell(5, 2)
        cell.value = f'売上ランキング'
        cell.font = Font(bold=True, color='008080', size=16)

        cell = ws.cell(5, 5)
        cell.value = f'{rank}位'
        cell.font = Font(bold=True, color='008080', size=16)

        cell = ws.cell(6, 2)
        cell.value = f'売上データ'
        cell.font = Font(bold=True, color='008080', size=16)

        # 表の貼り付け
        tmp_df = df.loc[(df['store_id'] == target_id)
                        & (df['status'].isin([1, 2]))]
        tmp_df = tmp_df[[
            'order_accept_date',
            'customer_id',
            'total_amount',
            'takeout_name',
            'status_name']]
        data_export(tmp_df, ws, 7, 2)

        # キャンセル率ランキングを直接出力
        cell = ws. cell(5, 8)
        cell.value = f'キャンセル率ランキング'
        cell.font = Font(bold=True, color='008080', size=16)

        cell = ws.cell(5, 12)
        cell.value = f'{cancel_rank}位 {cancel_count.values[0]}回'
        cell.font = Font(bold=True, color='008080', size=16)

        cell = ws. cell(6, 8)
        cell.value = f'キャンセルデータ'
        cell.font = Font(bold=True, color='008080', size=16)

        # 表の貼り付け
        tmp_df = df.loc[(df['store_id'] == target_id) & (df['status'] == 9)]
        tmp_df = tmp_df[[
            'order_accept_date',
            'customer_id',
            'total_amount',
            'takeout_name',
            'status_name']]
        data_export(tmp_df, ws, 7, 8)

        # 配達完了までの時間を直接出力
        ave_time = delivery_df.loc[
            delivery_df['store_id'] == target_id]['delta'].values[0]

        cell = ws.cell(5, 14)
        cell.value = f'配達完了までの時間ランキング'
        cell.font = Font(bold=True, color='008080', size=16)

        cell = ws.cell(5, 18)
        cell.value = f'{delivery_rank}位 平均{ave_time}分'
        cell.font = Font(bold=True, color='008080', size=16)

        cell = ws.cell(6, 14)
        cell.value = f'各店舗の配達時間ランク'
        cell.font = Font(bold=True, color='008080', size=16)

        # 表の貼り付け
        data_export(delivery_df, ws, 7, 14)

    # デフォルトシートの削除
    wb.remove(wb.worksheets[0])

    # DFループが終わったらブックを保存
    wb.save(os.path.join(output_folder,
            f'{target_id}_{store_name}_report_{file_date}.xlsx'))
    wb.close()


# %%
